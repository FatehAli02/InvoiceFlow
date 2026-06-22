import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from core.models import Invoice
from utils.loggers import logger

def generate_excel_report(invoices : list[Invoice], output_path : Path):

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    raw_sheet = wb.active
    raw_sheet.title = "Raw Data"

    raw_sheet.append(["Invoice No.", "Vendor", "Date", "Amount"])

    for cell in raw_sheet[1]:  
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    raw_sheet.column_dimensions['A'].width = 15
    raw_sheet.column_dimensions['B'].width = 30
    raw_sheet.column_dimensions['C'].width = 15
    raw_sheet.column_dimensions['D'].width = 15

    total_invoices = len(invoices)
    total_amount = 0.0
    vendors_total = {}

    for inv in invoices:

        raw_sheet.append([inv.invoice_number, inv.vendor, inv.date, inv.total_amount])

        total_amount += inv.total_amount

        if inv.vendor in vendors_total:
            vendors_total[inv.vendor] += inv.total_amount
        else:
            vendors_total[inv.vendor] = inv.total_amount
    
    summary_sheet = wb.create_sheet(title="Summary")

    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 20

    summary_sheet.append(["Overall Summary:", ""])

    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    summary_sheet.append(["Total Invoices:", total_invoices])
    summary_sheet.append(["Total Amount:", total_amount])
    summary_sheet.append([])

    summary_sheet.append(["Vendor-wise Totals", "Total Amount"])

    for cell in summary_sheet[5]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    summary_sheet.append(["Vendor", "Total"])
    for vendor, amount in vendors_total.items():
        summary_sheet.append([vendor, amount])
    
    filename = "invoice_report.xlsx"

    file_path = output_path/filename
    output_path.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    logger.info(f"Excel report successfully saved to: {output_path}")